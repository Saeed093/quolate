"""M5 sell-side quotation tests. LLM mocked; pricing/versioning is deterministic."""
from __future__ import annotations

import io
import json

from app.llm.mock import queue_responses


# ---------- helpers ----------
def _project(auth_client, **fields) -> str:
    pid = auth_client.post("/projects", json={"name": "Acme RFP"}).json()["id"]
    if fields:
        auth_client.patch(f"/projects/{pid}", json=fields)
    return pid


def _add_bom(auth_client, pid: str, part_name: str, quantity=None) -> str:
    body = {"part_name": part_name}
    if quantity is not None:
        body["quantity"] = quantity
    return auth_client.post(f"/projects/{pid}/bom", json=body).json()["id"]


def _create_quotation(auth_client, pid: str) -> dict:
    resp = auth_client.post(f"/projects/{pid}/quotations", json={"sources": []})
    assert resp.status_code == 201, resp.text
    return resp.json()


# ---------- extraction ----------
def test_extract_requirements_from_text(auth_client):
    pid = _project(auth_client)
    queue_responses(
        json.dumps(
            {
                "line_items": [
                    {"part_name": "Office chair", "spec_requirement": "mesh", "quantity": 50},
                    {"part_name": "Filing cabinet", "spec_requirement": None, "quantity": 10},
                ]
            }
        )
    )
    resp = auth_client.post(
        f"/projects/{pid}/quotations/extract-requirements",
        json={"sources": [{"kind": "text", "text": "Need 50 mesh chairs and 10 cabinets"}]},
    )
    assert resp.status_code == 201, resp.text
    items = resp.json()
    assert [i["part_name"] for i in items] == ["Office chair", "Filing cabinet"]
    # Persisted to the BOM.
    bom = auth_client.get(f"/projects/{pid}/bom").json()
    assert len(bom) == 2


# ---------- assembly + gaps ----------
def test_assembly_flags_gap_when_no_supplier_cost(auth_client):
    pid = _project(auth_client)
    _add_bom(auth_client, pid, "Exotic widget", quantity=3)
    q = _create_quotation(auth_client, pid)
    version = q["versions"][0]
    line = version["lines"][0]
    assert line["gap_flag"] is True
    assert line["unit_cost"] is None
    assert line["unit_price"] is None
    assert float(version["subtotal"]) == 0.0


def test_quote_numbering_is_per_project_sequence(auth_client):
    pid = _project(auth_client)
    _add_bom(auth_client, pid, "Item")
    first = _create_quotation(auth_client, pid)["quote_no"]
    second = _create_quotation(auth_client, pid)["quote_no"]
    assert first == "acme-rfp-QUO-0001"
    assert second == "acme-rfp-QUO-0002"


# ---------- margin + rounding + GST ----------
def test_margin_applied_and_rounded_half_up(auth_client):
    pid = _project(auth_client)
    lid = _add_bom(auth_client, pid, "Widget", quantity=1)
    q = _create_quotation(auth_client, pid)
    version = q["versions"][0]
    line_id = version["lines"][0]["id"]
    # margin 12.5% on a cost of 100 => 112.5 -> half-up -> 113 (banker's would give 112).
    resp = auth_client.patch(
        f"/projects/{pid}/quotations/versions/{version['id']}",
        json={"margin_pct": 0.125, "lines": [{"id": line_id, "unit_cost": 100}]},
    )
    assert resp.status_code == 200, resp.text
    ln = resp.json()["lines"][0]
    assert float(ln["unit_price"]) == 113.0
    assert ln["gap_flag"] is False


def test_manual_price_resolves_gap_with_gst_totals(auth_client):
    pid = _project(auth_client, gst_enabled=True, gst_pct=0.18)
    lid = _add_bom(auth_client, pid, "Chair", quantity=2)
    q = _create_quotation(auth_client, pid)
    version = q["versions"][0]
    line_id = version["lines"][0]["id"]
    resp = auth_client.patch(
        f"/projects/{pid}/quotations/versions/{version['id']}",
        json={"lines": [{"id": line_id, "unit_price": 200}]},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    ln = body["lines"][0]
    assert ln["cost_source"] == "manual"
    assert float(ln["line_total"]) == 400.0  # 200 * qty 2
    assert float(body["subtotal"]) == 400.0
    assert float(body["tax_total"]) == 72.0  # 400 * 0.18
    assert float(body["grand_total"]) == 472.0


def test_remove_line_updates_totals(auth_client):
    pid = _project(auth_client)
    _add_bom(auth_client, pid, "A", quantity=1)
    _add_bom(auth_client, pid, "B", quantity=1)
    q = _create_quotation(auth_client, pid)
    version = q["versions"][0]
    keep, drop = version["lines"][0]["id"], version["lines"][1]["id"]
    resp = auth_client.patch(
        f"/projects/{pid}/quotations/versions/{version['id']}",
        json={
            "lines": [
                {"id": keep, "unit_price": 50},
                {"id": drop, "remove": True},
            ]
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert len(body["lines"]) == 1
    assert float(body["subtotal"]) == 50.0


# ---------- versioning ----------
def test_finalize_makes_version_immutable(auth_client):
    pid = _project(auth_client)
    _add_bom(auth_client, pid, "Item")
    q = _create_quotation(auth_client, pid)
    vid = q["versions"][0]["id"]
    fin = auth_client.post(f"/projects/{pid}/quotations/versions/{vid}/finalize")
    assert fin.status_code == 200
    assert fin.json()["status"] == "final"
    # Editing a finalized version is rejected.
    resp = auth_client.patch(
        f"/projects/{pid}/quotations/versions/{vid}", json={"margin_pct": 0.5}
    )
    assert resp.status_code == 409


def test_regenerate_creates_new_version_retaining_old(auth_client):
    pid = _project(auth_client)
    lid = _add_bom(auth_client, pid, "Item", quantity=1)
    q = _create_quotation(auth_client, pid)
    v1 = q["versions"][0]
    # Put a manual price on v1 so we can confirm it carries into v2.
    auth_client.patch(
        f"/projects/{pid}/quotations/versions/{v1['id']}",
        json={"lines": [{"id": v1["lines"][0]["id"], "unit_price": 77}]},
    )
    auth_client.post(f"/projects/{pid}/quotations/versions/{v1['id']}/finalize")
    regen = auth_client.post(
        f"/projects/{pid}/quotations/versions/{v1['id']}/regenerate"
    )
    assert regen.status_code == 200
    v2 = regen.json()
    assert v2["version_no"] == 2
    assert v2["status"] == "draft"
    assert float(v2["lines"][0]["unit_price"]) == 77.0  # cloned
    # Both versions retained.
    full = auth_client.get(f"/projects/{pid}/quotations/{q['id']}").json()
    assert len(full["versions"]) == 2
    assert full["versions"][0]["status"] == "final"


# ---------- render / download ----------
def test_download_docx_and_xlsx(auth_client):
    pid = _project(auth_client, gst_enabled=True, gst_pct=0.18)
    lid = _add_bom(auth_client, pid, "Chair", quantity=2)
    q = _create_quotation(auth_client, pid)
    version = q["versions"][0]
    auth_client.patch(
        f"/projects/{pid}/quotations/versions/{version['id']}",
        json={"lines": [{"id": version["lines"][0]["id"], "unit_price": 200}]},
    )

    docx = auth_client.get(
        f"/projects/{pid}/quotations/versions/{version['id']}/download?fmt=docx"
    )
    assert docx.status_code == 200
    assert docx.content[:2] == b"PK"  # zip/ooxml magic
    from docx import Document

    Document(io.BytesIO(docx.content))  # opens without error

    xlsx = auth_client.get(
        f"/projects/{pid}/quotations/versions/{version['id']}/download?fmt=xlsx"
    )
    assert xlsx.status_code == 200
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx.content))
    assert "Buildup" in wb.sheetnames and "Totals" in wb.sheetnames

    bad = auth_client.get(
        f"/projects/{pid}/quotations/versions/{version['id']}/download?fmt=pdf"
    )
    assert bad.status_code == 400
