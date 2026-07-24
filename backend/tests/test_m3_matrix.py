"""M3 matrix + landed-cost tests. LLM mocked; math is deterministic."""
from __future__ import annotations

import asyncio
import io
import json
from decimal import Decimal

from app.llm.mock import queue_responses
from app.matrix.fx import convert
from app.matrix.landed_cost import landed_unit_cost


def _drain() -> int:
    from app.jobs.worker import drain

    return asyncio.run(drain())


# ---------- Pure math ----------
def test_landed_cost_formula_exact_values():
    # 100 * (1 + 0.10 + 0.05) + 3 = 118
    result = landed_unit_cost(
        Decimal("100"),
        duty_pct=Decimal("0.10"),
        lc_pct=Decimal("0.05"),
        freight_per_unit=Decimal("3"),
    )
    assert result == Decimal("118.00")

    # No add-ons -> equals FOB.
    assert landed_unit_cost(Decimal("50")) == Decimal("50")

    # Coerces plain numbers/strings.
    assert landed_unit_cost("10", duty_pct=0.2, freight_per_unit=1, lc_pct=0) == Decimal(
        "13.0"
    )


def test_currency_conversion_with_manual_override():
    # Override CNY to exactly 7.0 per USD -> 700 CNY == 100 USD.
    usd = convert(Decimal("700"), "CNY", "USD", {"CNY": 7.0})
    assert usd == Decimal("100")

    # Same currency is a no-op.
    assert convert(Decimal("42"), "USD", "USD") == Decimal("42")

    # Bundled static rate still works without an override (EUR 0.92 per USD).
    eur = convert(Decimal("92"), "EUR", "USD")
    assert eur == Decimal("100")


# ---------- Integration helpers ----------
def _project(auth_client, defaults: dict | None = None) -> str:
    body = {"name": "P"}
    if defaults is not None:
        body["landed_cost_defaults"] = defaults
    return auth_client.post("/projects", json=body).json()["id"]


def _add_bom(auth_client, pid: str, part: str) -> int:
    return auth_client.post(f"/projects/{pid}/bom", json={"part_name": part}).json()[
        "line_no"
    ]


def _upload_quote(
    auth_client,
    pid: str,
    *,
    supplier: str,
    line_no: int,
    price: float,
    confidence: float = 0.95,
    currency: str = "USD",
    extra_fields: list[dict] | None = None,
) -> None:
    queue_responses(
        json.dumps(
            {
                "supplier_name": supplier,
                "currency": currency,
                "fields": [
                    {
                        "bom_line_no": line_no,
                        "field_type": "unit_price",
                        "value_num": price,
                        "currency": currency,
                        "confidence": confidence,
                        "source_snippet": f"{supplier} {part_price(line_no, price)}",
                    },
                    *(extra_fields or []),
                ],
            }
        )
    )
    fn = f"{supplier}-{line_no}-{price}.txt"
    content = f"{supplier} line {line_no} price {price}".encode()
    resp = auth_client.post(
        f"/projects/{pid}/documents",
        files=[("files", (fn, content, "text/plain"))],
    )
    assert resp.status_code == 201
    assert _drain() >= 1


def part_price(line_no: int, price: float) -> str:
    return f"L{line_no} {price}"


# ---------- Matrix builder ----------
def test_best_value_flag_lowest_landed(auth_client):
    pid = _project(auth_client)
    line = _add_bom(auth_client, pid, "Widget A")
    _upload_quote(auth_client, pid, supplier="ACME", line_no=line, price=12.5)
    _upload_quote(auth_client, pid, supplier="Globex", line_no=line, price=10.0)

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    row = matrix["rows"][0]
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}

    globex_cell = row["cells"][suppliers["Globex"]]
    acme_cell = row["cells"][suppliers["ACME"]]
    assert globex_cell["best_value"] is True
    assert acme_cell["best_value"] is False
    assert row["best_supplier_id"] == suppliers["Globex"]


def test_selected_supplier_overrides_best_in_quotation(auth_client):
    pid = _project(
        auth_client, defaults={"duty_pct": 0, "lc_pct": 0, "freight_per_unit": 0}
    )
    line = _add_bom(auth_client, pid, "Widget A")
    _upload_quote(auth_client, pid, supplier="Cheap", line_no=line, price=100.0)
    _upload_quote(auth_client, pid, supplier="Pricey", line_no=line, price=150.0)

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    row = matrix["rows"][0]
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}
    assert row["best_supplier_id"] == suppliers["Cheap"]
    assert row["selected_supplier_id"] is None

    # By default the quotation uses the cheapest covering supplier (100).
    q1 = auth_client.post(f"/projects/{pid}/quotations", json={"sources": []}).json()
    assert float(q1["versions"][0]["lines"][0]["unit_cost"]) == 100.0

    # Pick the pricier supplier for this line.
    bom_item_id = row["bom_item_id"]
    r = auth_client.patch(
        f"/bom/{bom_item_id}", json={"selected_supplier_id": suppliers["Pricey"]}
    )
    assert r.status_code == 200, r.text
    assert r.json()["selected_supplier_id"] == suppliers["Pricey"]
    matrix2 = auth_client.get(f"/projects/{pid}/matrix").json()
    assert matrix2["rows"][0]["selected_supplier_id"] == suppliers["Pricey"]

    # A new quotation now uses the selected (pricier) supplier's landed cost.
    q2 = auth_client.post(f"/projects/{pid}/quotations", json={"sources": []}).json()
    assert float(q2["versions"][0]["lines"][0]["unit_cost"]) == 150.0

    # A supplier from another project (or bogus id) is rejected.
    bad = auth_client.patch(
        f"/bom/{bom_item_id}",
        json={"selected_supplier_id": "00000000-0000-0000-0000-000000000000"},
    )
    assert bad.status_code == 422

    # Clearing the selection reverts the quote to the cheapest.
    auth_client.patch(f"/bom/{bom_item_id}", json={"selected_supplier_id": None})
    q3 = auth_client.post(f"/projects/{pid}/quotations", json={"sources": []}).json()
    assert float(q3["versions"][0]["lines"][0]["unit_cost"]) == 100.0


def test_gap_cell_when_supplier_missing_line(auth_client):
    pid = _project(auth_client)
    l1 = _add_bom(auth_client, pid, "Widget A")
    l2 = _add_bom(auth_client, pid, "Widget B")
    # ACME quotes both lines; Globex only line 1.
    _upload_quote(auth_client, pid, supplier="ACME", line_no=l1, price=12.5)
    _upload_quote(auth_client, pid, supplier="ACME", line_no=l2, price=8.0)
    _upload_quote(auth_client, pid, supplier="Globex", line_no=l1, price=10.0)

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}
    line2_row = next(r for r in matrix["rows"] if r["line_no"] == l2)
    globex_cell = line2_row["cells"][suppliers["Globex"]]
    assert globex_cell["confidence_state"] == "gap"
    assert globex_cell["landed"] is None


def test_verify_state_when_any_field_unconfirmed_below_threshold(auth_client):
    pid = _project(auth_client)
    line = _add_bom(auth_client, pid, "Widget A")
    _upload_quote(
        auth_client, pid, supplier="ACME", line_no=line, price=12.5, confidence=0.4
    )

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}
    cell = matrix["rows"][0]["cells"][suppliers["ACME"]]
    assert cell["confidence_state"] == "verify"

    # Confirming the field flips the cell to ok.
    review = auth_client.get(
        f"/documents/{_first_doc(auth_client, pid)}/review"
    ).json()
    field_id = next(f["id"] for f in review["fields"] if f["field_type"] == "unit_price")
    auth_client.patch(f"/fields/{field_id}", json={"status": "confirmed"})

    matrix2 = auth_client.get(f"/projects/{pid}/matrix").json()
    cell2 = matrix2["rows"][0]["cells"][suppliers["ACME"]]
    assert cell2["confidence_state"] == "ok"


def test_export_xlsx_opens_and_matches_matrix(auth_client):
    from openpyxl import load_workbook

    pid = _project(auth_client, defaults={"duty_pct": 0.1, "lc_pct": 0.05, "freight_per_unit": 3})
    line = _add_bom(auth_client, pid, "Widget A")
    _upload_quote(auth_client, pid, supplier="ACME", line_no=line, price=100.0)

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}
    landed = matrix["rows"][0]["cells"][suppliers["ACME"]]["landed"]
    assert landed == 118.0  # 100 * 1.15 + 3

    resp = auth_client.get(f"/projects/{pid}/matrix/export")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb["Matrix"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Line"
    acme_col = next(i for i, h in enumerate(header) if h and "ACME" in str(h))
    # Final column names the supplier chosen for the quotation.
    assert header[-1] == "Selected supplier"
    data_row = [c.value for c in ws[2]]
    assert data_row[acme_col] == 118.0
    # ACME is the only (hence best/selected) supplier for the line.
    assert data_row[-1] == "ACME"


def _first_doc(auth_client, pid: str) -> str:
    docs = auth_client.get(f"/projects/{pid}/documents").json()
    return docs[0]["id"]


# ---------- Document-level commercial terms ----------
def test_doc_level_terms_surface_in_every_cell(auth_client):
    pid = _project(auth_client)
    l1 = _add_bom(auth_client, pid, "Widget A")
    l2 = _add_bom(auth_client, pid, "Widget B")
    _upload_quote(
        auth_client,
        pid,
        supplier="ACME",
        line_no=l1,
        price=12.5,
        extra_fields=[
            {
                "bom_line_no": l2,
                "field_type": "unit_price",
                "value_num": 8.0,
                "confidence": 0.95,
                "source_snippet": "L2 8.0",
            },
            # Quotation-wide terms: bom_line_no null.
            {
                "bom_line_no": None,
                "field_type": "incoterms",
                "value_text": "FOB",
                "confidence": 0.9,
                "source_snippet": "Terms: FOB Shenzhen",
            },
            {
                "bom_line_no": None,
                "field_type": "lead_time_days",
                "value_num": 30,
                "confidence": 0.9,
                "source_snippet": "Lead time: 30 days",
            },
            {
                "bom_line_no": None,
                "field_type": "payment_terms",
                "value_text": "T/T 30% deposit",
                "confidence": 0.85,
                "source_snippet": "Payment: T/T 30% deposit",
            },
        ],
    )

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}
    for row in matrix["rows"]:
        cell = row["cells"][suppliers["ACME"]]
        assert cell["incoterms"] == "FOB"
        assert cell["lead_time_days"] == 30
        assert cell["payment_terms"] == "T/T 30% deposit"


def test_line_level_term_wins_over_doc_level(auth_client):
    pid = _project(auth_client)
    line = _add_bom(auth_client, pid, "Widget A")
    _upload_quote(
        auth_client,
        pid,
        supplier="ACME",
        line_no=line,
        price=12.5,
        extra_fields=[
            {
                "bom_line_no": line,
                "field_type": "incoterms",
                "value_text": "CIF",
                "confidence": 0.9,
                "source_snippet": "Widget A CIF Karachi",
            },
            {
                "bom_line_no": None,
                "field_type": "incoterms",
                "value_text": "FOB",
                "confidence": 0.9,
                "source_snippet": "Other terms FOB",
            },
        ],
    )

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}
    cell = matrix["rows"][0]["cells"][suppliers["ACME"]]
    assert cell["incoterms"] == "CIF"


def test_doc_level_term_not_fuzzy_matched_onto_one_line(auth_client):
    """A doc-level term whose snippet mentions a part name must stay doc-level."""
    pid = _project(auth_client)
    l1 = _add_bom(auth_client, pid, "Thermal Camera Module")
    l2 = _add_bom(auth_client, pid, "Thermal Lens Assembly")
    _upload_quote(
        auth_client,
        pid,
        supplier="ACME",
        line_no=l1,
        price=132.0,
        extra_fields=[
            {
                "bom_line_no": l2,
                "field_type": "unit_price",
                "value_num": 44.0,
                "confidence": 0.95,
                "source_snippet": "L2 44.0",
            },
            {
                "bom_line_no": None,
                "field_type": "incoterms",
                "value_text": "EXW",
                # Snippet mentions a part name — must NOT bucket to line 1 only.
                "source_snippet": "All items incl. Thermal Camera Module ship EXW",
                "confidence": 0.9,
            },
        ],
    )

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}
    for row in matrix["rows"]:
        assert row["cells"][suppliers["ACME"]]["incoterms"] == "EXW"


def test_terms_only_document_still_creates_quote(auth_client):
    """An incoterms-only extraction (no price) must still surface in the matrix."""
    pid = _project(auth_client)
    _add_bom(auth_client, pid, "Widget A")
    queue_responses(
        json.dumps(
            {
                "supplier_name": "TermsOnly Co",
                "currency": "USD",
                "fields": [
                    {
                        "bom_line_no": None,
                        "field_type": "incoterms",
                        "value_text": "DDP",
                        "confidence": 0.9,
                        "source_snippet": "All prices DDP",
                    }
                ],
            }
        )
    )
    resp = auth_client.post(
        f"/projects/{pid}/documents",
        files=[("files", ("terms.txt", b"All prices DDP", "text/plain"))],
    )
    assert resp.status_code == 201
    assert _drain() >= 1

    matrix = auth_client.get(f"/projects/{pid}/matrix").json()
    suppliers = {s["name"]: s["id"] for s in matrix["suppliers"]}
    cell = matrix["rows"][0]["cells"][suppliers["TermsOnly Co"]]
    assert cell["confidence_state"] == "gap"  # no price -> still a gap
    assert cell["incoterms"] == "DDP"
    assert cell["document_id"] is not None  # Open source works from doc-level quote
