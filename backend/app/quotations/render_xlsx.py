"""Render a quotation version's internal calculation buildup to XLSX.

Follows the pattern of app.matrix.export.matrix_to_xlsx. Unlike the client DOCX,
this sheet exposes the full working: landed unit cost, the margin applied, the
resulting sell price, and the cost source per line — for the user's own records.
"""
from __future__ import annotations

import io


def _num(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def render_quotation_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Buildup"

    ccy = data.get("currency", "")
    margin = _num(data.get("margin_pct"))
    margin_pct = round(margin * 100, 2) if margin is not None else None

    header = [
        "Line",
        "Description",
        "Spec",
        "Qty",
        f"Unit cost ({ccy})",
        "Margin %",
        f"Unit price ({ccy})",
        f"Line total ({ccy})",
        "Cost source",
        "Gap",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for ln in data.get("lines", []):
        ws.append(
            [
                ln.get("line_no"),
                ln.get("description"),
                ln.get("spec"),
                _num(ln.get("qty")),
                _num(ln.get("unit_cost")),
                margin_pct,
                _num(ln.get("unit_price")),
                _num(ln.get("line_total")),
                ln.get("cost_source"),
                "yes" if ln.get("gap_flag") else "",
            ]
        )

    ws2 = wb.create_sheet("Totals")
    ws2.append(["Metric", f"Value ({ccy})"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    ws2.append(["Quotation No", data.get("quote_no")])
    ws2.append(["Subtotal", _num(data.get("subtotal"))])
    if data.get("gst_enabled"):
        gst = _num(data.get("gst_pct"))
        label = f"GST ({round(gst * 100, 2)}%)" if gst is not None else "GST"
        ws2.append([label, _num(data.get("tax_total"))])
    ws2.append(["Grand total", _num(data.get("grand_total"))])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
