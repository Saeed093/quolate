"""Render a computed matrix to XLSX (openpyxl), server-side."""
from __future__ import annotations

import io


def matrix_to_xlsx(matrix: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Matrix"

    currency = matrix.get("currency", "USD")
    suppliers = matrix.get("suppliers", [])

    header = ["Line", "Part", "Spec", "Qty", "Target"]
    for s in suppliers:
        header.append(f"{s['name']} ({currency})")
    header.append("Selected supplier")
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    supplier_names = {s["id"]: s["name"] for s in suppliers}
    for row in matrix.get("rows", []):
        line = [
            row.get("line_no"),
            row.get("part_name"),
            row.get("spec_requirement"),
            row.get("quantity"),
            row.get("target_price"),
        ]
        cells = row.get("cells", {})
        for s in suppliers:
            cell = cells.get(s["id"], {})
            landed = cell.get("landed")
            line.append(landed if landed is not None else None)
        # Which supplier feeds the quotation: explicit pick (if priced) else best.
        selected = row.get("selected_supplier_id")
        best = row.get("best_supplier_id")
        chosen = (
            selected
            if selected and (cells.get(selected) or {}).get("landed") is not None
            else best
        )
        line.append(supplier_names.get(chosen, "") if chosen else "")
        ws.append(line)

    # Summary sheet.
    summary = matrix.get("summary", {})
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2["A1"].font = Font(bold=True)
    ws2["B1"].font = Font(bold=True)
    for key, value in summary.items():
        ws2.append([key, value])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
